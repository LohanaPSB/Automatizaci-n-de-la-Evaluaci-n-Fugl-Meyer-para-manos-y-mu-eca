from Leer_datos import leer_angulo
import tkinter as tk
from tkinter import messagebox
from se import mpu9250
import pandas as pd
import zipfile
import os
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from PIL import Image, ImageTk
import imageio
from numpy import interp

# Simulación de resultados de las evaluaciones (para propósitos de demostración)
# Cada evaluación tiene asociados dos resultados.
resultados_mano = {}
resultados_muneca = {}

def agregar_resultado_mano(evaluacion, resultado1, resultado2, resultado3):
    global resultados_mano
    resultados_mano[evaluacion] = (resultado1, resultado2, resultado3)

def agregar_resultado_muneca(evaluacion, resultado1, resultado2, resultado3):
    global resultados_muneca
    resultados_muneca[evaluacion] = (resultado1, resultado2, resultado3)

class VideoPlayer(tk.Label):
    def __init__(self, master, video_path, **kwargs):
        super().__init__(master, **kwargs)
        self.video_path = video_path
        self.video = imageio.get_reader(video_path)
        self.delay = int(500 / self.video.get_meta_data()['fps'])
        self.frame_gen = self.video.iter_data()
        self.playing = False

    def play(self):
        self.playing = True
        self.next_frame()

    def stop(self):
        self.playing = False

    def next_frame(self):
        if self.playing:
            try:
                frame = next(self.frame_gen)
                image = Image.fromarray(frame)
                photo = ImageTk.PhotoImage(image)
                self.config(image=photo)
                self.image = photo
                self.after(self.delay, self.next_frame)
            except StopIteration:
                self.stop()
# Función para cambiar a la ventana de evaluación FMA-Mano
# ... (El resto del código anterior permanece igual)

# Función para cambiar a la ventana de evaluación FMA-Mano
def abrir_fma_mano():
    ventana_fma_mano = tk.Toplevel()
    ventana_fma_mano.title("Evaluación FMA-Mano")

    # Creamos el widget Text y lo configuramos para que sea de solo lectura
    texto_descripcion = tk.Text(ventana_fma_mano, wrap=tk.WORD, height=16, width=50)
    texto_descripcion.grid(row=0, column=0, columnspan=2)

    # Insertamos la descripción en el widget Text
    descripcion = ("Se va a llevar a cabo la evaluación Fugl-Meyer de la mano,\n"
                   "esta consta de 7 movimientos:\n"
                   "1. Flexión de la mano\n"
                   "2. Extensión de la mano\n"
                   "3. Agarre de gancho\n"
                   "4. Oposición del pulgar\n"
                   "5. Agarre de pinza\n"
                   "6. Agarre cilíndrico\n"
                   "7. Agarre esférico\n\n"
                   "Cada uno de los movimientos se clasifica según la siguiente escala:\n"
                   "0 - no desarrolla el movimiento\n"
                   "1 - desarrolla el movimiento de forma parcial\n"
                   "2 - desarrolla el movimiento completamente")

    # Insertamos el texto en el widget Text
    texto_descripcion.insert(tk.END, descripcion)

    # Centramos el texto utilizando una etiqueta de formato
    texto_descripcion.tag_configure("center", justify="center")
    texto_descripcion.tag_add("center", "1.0", "end")

    # Hacemos que el widget Text sea de solo lectura
    texto_descripcion.config(state=tk.DISABLED)


    # Botones para cada movimiento, organizados en dos columnas y uno en el centro
    movimientos = [
        ("1. Flexión", 6, 0, abrir_flexion), ("2. Extensión", 6, 1, abrir_extension),
        ("3. Agarre de Gancho", 7, 0, abrir_agarre_gancho), ("4.Oposición del Pulgar", 7, 1, abrir_opo_pulgar),
        ("5. Agarre de Pinza", 8, 0, abrir_agarre_pinza), ("6. Agarre Cilíndrico", 8, 1, abrir_agarre_cilindrico),("6. Agarre Cilíndrico", 8, 1, abrir_agarre_cilindrico)
    ]

    button_width = 20  # Ancho deseado de los botones

    for movimiento, fila, columna, ventana in movimientos[:-1]:
        tk.Button(ventana_fma_mano, text=movimiento, command = ventana, width=button_width).grid(row=fila, column=columna, padx=5, pady=5)

    # Colocamos el botón 'Agarre Esférico' en el centro de la tercera fila
    tk.Button(ventana_fma_mano, text="7. Agarre Esférico", command= abrir_agarre_esferico, width=button_width).grid(row=9, column=0, columnspan=2, padx=5, pady=5)

    # Botón para volver a la ventana de inicio
    tk.Button(ventana_fma_mano, text="Volver", command=ventana_fma_mano.destroy, width=button_width).grid(row=10, column=0, columnspan=2, padx=5, pady=5)
######################### FLEXION DE LA MANO ###################################################################
    
def evaluar_flexion(x, mano, evaluacion):
    print(x)
    angulos = x.values()
    if mano == "izquierda":
        if all(-5 <= angulo <= 5 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,0, x)
            return 0
        elif all(5 <= angulo <= 50 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,1, x)
            return 1
        elif all( 50 <= angulo <= 100 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,2, x)
            return 2
        else:
            print(x)
            return None
        pass
    elif mano == "derecha":
        # Logica para evaluar la mano derecha
        if all(-5 <= angulo <= 5 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,0, x)
            return 0
        elif all(5<= angulo <= 50 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,1, x)
            return 1
        elif all( 50 <= angulo <= 100 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,2, x)
            return 2
        else:
            print(x)
            return None
        pass


def verificar_posicion_inicial(mano):
    valores = leer_angulo(mano)
    print(valores)
    if all(-3 <= valor <= 3 for valor in valores.values()):
        
        return True
    return False

# Nueva función para realizar la evaluación de flexión después de la espera
def realizar_evaluacion(mano, evaluacion):
    resultado = evaluar_flexion(leer_angulo(mano), mano, evaluacion)
    if resultado is not None:
        resultado_label.config(text=f"Resultado de la flexión: {resultado}")
    else:
        resultado_label.config(text="Valores fuera de rango.")

def iniciar_evaluacion_flexion(mano,ventana, evaluacion):
    if verificar_posicion_inicial(mano):
        resultado_label.config(text="Realice el movimiento de Flexión.")
        # Esperar 3 segundos antes de realizar la evaluación
        ventana.after(5000, lambda: realizar_evaluacion(mano, evaluacion))
    else:
        resultado_label.config(text="Por favor, ajusta la posición inicial.")

def abrir_flexion():
    ventana_flexion = tk.Toplevel()
    ventana_flexion.title("Flexión de la Mano")

    
    tk.Label(ventana_flexion, text="Selecciona la mano:").pack()
    mano_var = tk.StringVar(value="izquierda")
    tk.Radiobutton(ventana_flexion, text="Izquierda", variable=mano_var, value="izquierda").pack()
    tk.Radiobutton(ventana_flexion, text="Derecha", variable=mano_var, value="derecha").pack()

    tk.Label(ventana_flexion, text="Nombre de la evaluación:").pack()
    entrada_evaluacion = tk.Entry(ventana_flexion)
    entrada_evaluacion.pack()
    
    iniciar_btn = tk.Button(ventana_flexion, text="Iniciar Evaluación",
                            command=lambda: iniciar_evaluacion_flexion(mano_var.get(), ventana_flexion, entrada_evaluacion.get()))
    iniciar_btn.pack()
    
    global resultado_label
    resultado_label = tk.Label(ventana_flexion, text="")
    resultado_label.pack()

    
    video_label = tk.Label(ventana_flexion, text="Video de referencia:")
    video_label.pack()
    
    videoplayer = VideoPlayer(ventana_flexion, "flexmano.mp4")
    videoplayer.pack(expand=True, fill="both")
    videoplayer.play()
########################################## EXTENSION DE LA MANO #####################################################
def evaluar_extension(x, mano, evaluacion):
    print(x)
    
    flex3 = x.pop('Flex3')
    flex4 = x.pop('Flex4')
    if mano == "izquierda":
        if 60 <= flex3 <= 100 and 60 <= flex4 <= 100 :
            a = {'Flex3': flex3,'Flex4':flex4}
            agregar_resultado_mano(evaluacion, mano,0, a)
            return 0
        elif  3 <= flex3 <= 60 and 3 <= flex4 <= 60:
            a = {'Flex3': flex3,'Flex4':flex4}
            agregar_resultado_mano(evaluacion, mano,1, a)
            return 1
        elif  -30 <= flex3 <= 3 and -30 <= flex4 <= 3:
            a = {'Flex3': flex3,'Flex4':flex4}
            agregar_resultado_mano(evaluacion, mano,2, a)
            return 2
        else:
            print(x)
            return None
        pass
    elif mano == "derecha":
        # Logica para evaluar la mano derecha
        if 60 <= flex3 <= 100 and 60 <= flex4 <= 100 :
            a = {'Flex3': flex3,'Flex4':flex4}
            agregar_resultado_mano(evaluacion, mano,0, a)
            return 0
        elif  3 <= flex3 <= 60 and 3 <= flex4 <= 60:
            a = {'Flex3': flex3,'Flex4':flex4}
            agregar_resultado_mano(evaluacion, mano,1, a)
            return 1
        elif  -30 <= flex3 <= 3 and -30 <= flex4 <= 3:
            a = {'Flex3': flex3,'Flex4':flex4}
            agregar_resultado_mano(evaluacion, mano,2, a)
            return 2
        else:
            print(x)
            return None
        pass


def verificar_posicion_iniciale(mano):
    valores = leer_angulo(mano)
    print(valores)
    if all(50 <= valor <= 100 for valor in valores.values()):
        return True
    return False

# Nueva función para realizar la evaluación de flexión después de la espera
def realizar_evaluacione(mano, evaluacion):
    resultado = evaluar_extension(leer_angulo(mano), mano, evaluacion)
    if resultado is not None:
        resultado_label.config(text=f"Resultado de la Extensión: {resultado}")
    else:
        resultado_label.config(text="Valores fuera de rango.")

def iniciar_evaluacion_extension(mano,ventana, evaluacion):
    if verificar_posicion_iniciale(mano):
        resultado_label.config(text="Realice el movimiento de Extensión.")
        # Esperar 3 segundos antes de realizar la evaluación
        ventana.after(5000, lambda: realizar_evaluacione(mano, evaluacion))
    else:
        resultado_label.config(text="Por favor, ajusta la posición inicial.")

def abrir_extension():
    ventana_extension = tk.Toplevel()
    ventana_extension.title("Extensión de la Mano")
     
    tk.Label(ventana_extension, text="Selecciona la mano:").pack()
    mano_var = tk.StringVar(value="izquierda")
    tk.Radiobutton(ventana_extension, text="Izquierda", variable=mano_var, value="izquierda").pack()
    tk.Radiobutton(ventana_extension, text="Derecha", variable=mano_var, value="derecha").pack()

    tk.Label(ventana_extension, text="Nombre de la evaluación:").pack()
    entrada_evaluacion = tk.Entry(ventana_extension)
    entrada_evaluacion.pack()
    
    iniciar_btn = tk.Button(ventana_extension, text="Iniciar Evaluación",
                            command=lambda: iniciar_evaluacion_extension(mano_var.get(), ventana_extension, entrada_evaluacion.get()))
    iniciar_btn.pack()
    
    global resultado_label
    resultado_label = tk.Label(ventana_extension, text="")
    resultado_label.pack()

    video_label = tk.Label(ventana_extension, text="Video de referencia:")
    video_label.pack()
    
    videoplayer = VideoPlayer(ventana_extension, "exmano.mp4")
    videoplayer.pack(expand=True, fill="both")
    videoplayer.play()

####################################### AGARRE GANCHO ##############################################################
def evaluar_gancho(x, mano, evaluacion):
    if mano == "izquierda":
        x.pop('PI')
        print(x)
        angulos = x.values()
        if all(-3 <= angulo <= 3 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,0, x)
            return 0
        elif all(3 <= angulo <= 60 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,1, x)
            return 1
        elif all( 60 <= angulo <= 100 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,2, x)
            return 2
        else:
            print(x)
            return None
        pass
    elif mano == "derecha":
        x.pop('PD')
        print(x)
        angulos = x.values()
        # Logica para evaluar la mano derecha
        if all(-3 <= angulo <= 3 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,0, x)
            return 0
        elif all(3<= angulo <= 60 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,1, x)
            return 1
        elif all( 60 <= angulo <= 100 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,2, x)
            return 2
        else:
            print(x)
            return None
        pass


def verificar_posicion_inicialg(mano):
    valores = leer_angulo(mano)
    print(valores)
    if mano == "izquierda":
        valores.pop('PI')
        if all(-3 <= valor <= 3 for valor in valores.values()):
            return True
        return False
    else: 
        valores.pop('PD')
        if all(-3 <= valor <= 3 for valor in valores.values()):
            return True
        return False
    
# Nueva función para realizar la evaluación de flexión después de la espera
def realizar_evaluaciong(mano, evaluacion):

    resultado = evaluar_gancho(leer_angulo(mano), mano, evaluacion)
    
    if resultado is not None:
        resultado_label.config(text=f"Resultado del Agarre de Gancho: {resultado}")
    else:
        resultado_label.config(text="Valores fuera de rango.")

def iniciar_evaluacion_gancho(mano,ventana, evaluacion):
    if verificar_posicion_inicialg(mano):
        resultado_label.config(text="Realice el movimiento de Agarre de Gancho.")
        # Esperar 3 segundos antes de realizar la evaluación
        ventana.after(5000, lambda: realizar_evaluaciong(mano, evaluacion))
    else:
        resultado_label.config(text="Por favor, ajusta la posición inicial.")


def abrir_agarre_gancho():
    ventana_gancho = tk.Toplevel()
    ventana_gancho.title("Agarre de Gancho")
    tk.Label(ventana_gancho, text="Selecciona la mano:").pack()
    mano_var = tk.StringVar(value="izquierda")
    tk.Radiobutton(ventana_gancho, text="Izquierda", variable=mano_var, value="izquierda").pack()
    tk.Radiobutton(ventana_gancho, text="Derecha", variable=mano_var, value="derecha").pack()

    tk.Label(ventana_gancho, text="Nombre de la evaluación:").pack()
    entrada_evaluacion = tk.Entry(ventana_gancho)
    entrada_evaluacion.pack()
    
    iniciar_btn = tk.Button(ventana_gancho, text="Iniciar Evaluación",
                            command=lambda: iniciar_evaluacion_gancho(mano_var.get(), ventana_gancho, entrada_evaluacion.get()))
    iniciar_btn.pack()
    
    global resultado_label
    resultado_label = tk.Label(ventana_gancho, text="")
    resultado_label.pack()

    video_label = tk.Label(ventana_gancho, text="Video de referencia:")
    video_label.pack()
    
    videoplayer = VideoPlayer(ventana_gancho, "gancho.mp4")
    videoplayer.pack(expand=True, fill="both")
    videoplayer.play()


###################################### OPOSICIÓN DEL PULGAR ########################################################
def evaluar_opo(x, mano, evaluacion):
    if mano == "izquierda":
        pulgar = x.pop('PI')
        print(x)
        angulos = x.values()
        if all(-3 <= angulo <= 3 for angulo in angulos):
            a = {'x': x,'PI':pulgar}
            agregar_resultado_mano(evaluacion, mano,0, a)
            return 0
        elif all(3 <= angulo <= 60 for angulo in angulos):
            a = {'x': x,'PI':pulgar}
            agregar_resultado_mano(evaluacion, mano,1, a)
            return 1
        elif all( 60 <= angulo <= 100 for angulo in angulos)and 0 <= pulgar <= 3:
            a = {'x': x,'PI':pulgar}
            agregar_resultado_mano(evaluacion, mano,2, a)
            return 2
        else:
            return None
        pass
    elif mano == "derecha":
        pulgar = x.pop('PD')
        print(x)
        angulos = x.values()
        # Logica para evaluar la mano derecha
        if all(-3 <= angulo <= 3 for angulo in angulos):
            a = {'x': x,'PD':pulgar}
            agregar_resultado_mano(evaluacion, mano,0, a)
            return 0
        elif all(3<= angulo <= 60 for angulo in angulos):
            a = {'x': x,'PD':pulgar}
            agregar_resultado_mano(evaluacion, mano,1, a)
            return 1
        elif all( 60 <= angulo <= 100 for angulo in angulos)and 0 <= pulgar <= 3:
            a = {'x': x,'PD':pulgar}
            agregar_resultado_mano(evaluacion, mano,2, a)
            return 2
        else:
            return None
        pass


def verificar_posicion_inicialo(mano):
    valores = leer_angulo(mano)
    print(valores)
    if all(-3 <= valor <= 3 for valor in valores.values()):
        return True
    return False

# Nueva función para realizar la evaluación de flexión después de la espera
def realizar_evaluaciono(mano, evaluacion):

    resultado = evaluar_opo(leer_angulo(mano), mano, evaluacion)
    
    if resultado is not None:
        resultado_label.config(text=f"Resultado de la Oposición del Pulgar: {resultado}")
    else:
        resultado_label.config(text="Valores fuera de rango.")

def iniciar_evaluacion_opo(mano,ventana, evaluacion):
    if verificar_posicion_inicialo(mano):
        resultado_label.config(text="Realice el movimiento de Oposición del Pulgar.")
        # Esperar 3 segundos antes de realizar la evaluación
        ventana.after(5000, lambda: realizar_evaluaciono(mano, evaluacion))
    else:
        resultado_label.config(text="Por favor, ajusta la posición inicial.")
def abrir_opo_pulgar():
    ventana_opo = tk.Toplevel()
    ventana_opo.title("Oposición de Pulgar")
    tk.Label(ventana_opo, text="Selecciona la mano:").pack()
    mano_var = tk.StringVar(value="izquierda")
    tk.Radiobutton(ventana_opo, text="Izquierda", variable=mano_var, value="izquierda").pack()
    tk.Radiobutton(ventana_opo, text="Derecha", variable=mano_var, value="derecha").pack()

    tk.Label(ventana_opo, text="Nombre de la evaluación:").pack()
    entrada_evaluacion = tk.Entry(ventana_opo)
    entrada_evaluacion.pack()
    
    iniciar_btn = tk.Button(ventana_opo, text="Iniciar Evaluación",
                            command=lambda: iniciar_evaluacion_opo(mano_var.get(), ventana_opo, entrada_evaluacion.get()))
    iniciar_btn.pack()
    
    global resultado_label
    resultado_label = tk.Label(ventana_opo, text="")
    resultado_label.pack()

    video_label = tk.Label(ventana_opo, text="Video de referencia:")
    video_label.pack()
    
    videoplayer = VideoPlayer(ventana_opo, "opo.mp4")
    videoplayer.pack(expand=True, fill="both")
    videoplayer.play()

###################################### AGARRE DE PINZA #############################################################
def evaluar_pinza(x, mano, evaluacion):
    if mano == "izquierda":
        pulgar = x.pop('PI')
        indice = x.pop('II')
        if -3 <= pulgar <= 3 and -3 <= indice <= 3:
            print(pulgar, indice)
            a = {'PI': pulgar, 'II': indice}
            agregar_resultado_mano(evaluacion, mano,0, a)
            return 0
        elif 3 <= pulgar <= 30 and 3 <= indice <= 30:
            print(pulgar, indice)
            a = {'PI': pulgar, 'II': indice}
            agregar_resultado_mano(evaluacion, mano,1, a)
            return 1
        elif 30 <= pulgar <= 60 and 30 <= indice <= 60:
            print(pulgar, indice)
            a = {'PI': pulgar, 'II': indice}
            agregar_resultado_mano(evaluacion, mano,2, a)
            return 2
        else:
            print(pulgar, indice)
            return None
        pass
    elif mano == "derecha":
        pulgar = x.pop('PD')
        indice = x.pop('ID')
        # Logica para evaluar la mano derecha
        if -3 <= pulgar <= 3 and -3 <= indice <= 3:
            print(pulgar, indice)
            a = {'PD': pulgar, 'ID': indice}
            agregar_resultado_mano(evaluacion, mano,0, a)
            return 0
        elif 3 <= pulgar <= 30 and 3 <= indice <= 30:
            print(pulgar, indice)
            a = {'PD': pulgar, 'ID': indice}
            agregar_resultado_mano(evaluacion, mano,1, a)
            return 1
        elif 30 <= pulgar <= 60 and 30 <= indice <= 60:
            print(pulgar, indice)
            a = {'PD': pulgar, 'ID': indice}
            agregar_resultado_mano(evaluacion, mano,0, a)
            return 2
        else:
            print(pulgar, indice)
            return None
        pass


def verificar_posicion_inicialp(mano):
    valores = leer_angulo(mano)
    print(valores)
    if all(-3 <= valor <= 3 for valor in valores.values()):
        return True
    return False

# Nueva función para realizar la evaluación de flexión después de la espera
def realizar_evaluacionp(mano, evaluacion):

    resultado = evaluar_pinza(leer_angulo(mano), mano, evaluacion)
    
    if resultado is not None:
        resultado_label.config(text=f"Resultado del Agarre de Pinza: {resultado}")
    else:
        resultado_label.config(text="Valores fuera de rango.")

def iniciar_evaluacion_pinza(mano,ventana, evaluacion):
    if verificar_posicion_inicialp(mano):
        resultado_label.config(text="Realice el movimiento de Agarre de Pinza.")
        # Esperar 3 segundos antes de realizar la evaluación
        ventana.after(5000, lambda: realizar_evaluacionp(mano, evaluacion))
    else:
        resultado_label.config(text="Por favor, ajusta la posición inicial.")
def abrir_agarre_pinza():
    ventana_pinza = tk.Toplevel()
    ventana_pinza.title("Agarre de Pinza")
    tk.Label(ventana_pinza, text="Selecciona la mano:").pack()
    mano_var = tk.StringVar(value="izquierda")
    tk.Radiobutton(ventana_pinza, text="Izquierda", variable=mano_var, value="izquierda").pack()
    tk.Radiobutton(ventana_pinza, text="Derecha", variable=mano_var, value="derecha").pack()

    tk.Label(ventana_pinza, text="Nombre de la evaluación:").pack()
    entrada_evaluacion = tk.Entry(ventana_pinza)
    entrada_evaluacion.pack()
    iniciar_btn = tk.Button(ventana_pinza, text="Iniciar Evaluación",
                            command=lambda: iniciar_evaluacion_pinza(mano_var.get(), ventana_pinza, entrada_evaluacion.get()))
    iniciar_btn.pack()
    
    global resultado_label
    resultado_label = tk.Label(ventana_pinza, text="")
    resultado_label.pack()

    video_label = tk.Label(ventana_pinza, text="Video de referencia:")
    video_label.pack()
    
    videoplayer = VideoPlayer(ventana_pinza, "pinza.mp4")
    videoplayer.pack(expand=True, fill="both")
    videoplayer.play()
###################################### AGARRE CILÍNDRICO ###########################################################
def evaluar_cilindrico(x, mano, evaluacion):
    if mano == "izquierda":
        pulgar = x.pop('PI')
        indice = x.pop('II')
        flex3 = x.pop('Flex3')
        if -3 <= pulgar <= 3 and -3 <= indice <= 3 and -3 <= flex3 <= 3:
            print(pulgar, indice, flex3)
            a = {'PI': pulgar,'II': indice, 'Flex3':flex3}
            agregar_resultado_mano(evaluacion, mano,0, a)
            return 0
        elif 3 <= pulgar <= 15 and 3 <= indice <= 15 and 3 <= flex3 <= 15:
            print(pulgar, indice, flex3)
            a = {'PI': pulgar,'II': indice, 'Flex3':flex3}
            agregar_resultado_mano(evaluacion, mano,1, a)
            return 1
        elif 15 <= pulgar <= 30 and 15 <= indice <= 30 and 15 <= flex3 <= 30:
            print(pulgar, indice, flex3)
            a = {'PI': pulgar,'II': indice, 'Flex3':flex3}
            agregar_resultado_mano(evaluacion, mano,2, a)
            return 2
        else:
            print(pulgar, indice, flex3)
            return None
        pass
    elif mano == "derecha":
        pulgar = x.pop('PD')
        indice = x.pop('ID')
        flex3 = x.pop('Flex3')
        if -3 <= pulgar <= 3 and -3 <= indice <= 3 and -3 <= flex3 <= 3:
            print(pulgar, indice, flex3)
            a = {'PD': pulgar,'ID': indice, 'Flex3':flex3}
            agregar_resultado_mano(evaluacion, mano,0, a)
            return 0
        elif 3 <= pulgar <= 15 and 3 <= indice <= 15 and 3 <= flex3 <= 15:
            print(pulgar, indice, flex3)
            a = {'PD': pulgar,'ID': indice, 'Flex3':flex3}
            agregar_resultado_mano(evaluacion, mano,1, a)
            return 1
        elif 15 <= pulgar <= 30 and 15 <= indice <= 30 and 15 <= flex3 <= 30:
            print(pulgar, indice, flex3)
            a = {'PD': pulgar,'ID': indice, 'Flex3':flex3}
            agregar_resultado_mano(evaluacion, mano,2, a)
            return 2
        else:
            print(pulgar, indice, flex3)
            return None


def verificar_posicion_inicialc(mano):
    valores = leer_angulo(mano)
    print(valores)
    if all(-3 <= valor <= 3 for valor in valores.values()):
        return True
    return False

# Nueva función para realizar la evaluación de flexión después de la espera
def realizar_evaluacionc(mano, evaluacion):

    resultado = evaluar_cilindrico(leer_angulo(mano), mano, evaluacion)
    
    if resultado is not None:
        resultado_label.config(text=f"Resultado del Agarre Cilíndrico: {resultado}")
    else:
        resultado_label.config(text="Valores fuera de rango.")

def iniciar_evaluacion_cilindrico(mano,ventana, evaluacion):
    if verificar_posicion_inicialc(mano):
        resultado_label.config(text="Realice el movimiento de Agarre Cilíndrico.")
        # Esperar 3 segundos antes de realizar la evaluación
        ventana.after(5000, lambda: realizar_evaluacionc(mano, evaluacion))
    else:
        resultado_label.config(text="Por favor, ajusta la posición inicial.")
def abrir_agarre_cilindrico():
    ventana_cilindrico = tk.Toplevel()
    ventana_cilindrico.title("Agarre Cilíndrico")
    tk.Label(ventana_cilindrico, text="Selecciona la mano:").pack()
    mano_var = tk.StringVar(value="izquierda")
    tk.Radiobutton(ventana_cilindrico, text="Izquierda", variable=mano_var, value="izquierda").pack()
    tk.Radiobutton(ventana_cilindrico, text="Derecha", variable=mano_var, value="derecha").pack()

    tk.Label(ventana_cilindrico, text="Nombre de la evaluación:").pack()
    entrada_evaluacion = tk.Entry(ventana_cilindrico)
    entrada_evaluacion.pack()   
    iniciar_btn = tk.Button(ventana_cilindrico, text="Iniciar Evaluación",
                            command=lambda: iniciar_evaluacion_cilindrico(mano_var.get(), ventana_cilindrico, entrada_evaluacion.get()))
    iniciar_btn.pack()
    
    global resultado_label
    resultado_label = tk.Label(ventana_cilindrico, text="")
    resultado_label.pack()

    video_label = tk.Label(ventana_cilindrico, text="Video de referencia:")
    video_label.pack()
    
    videoplayer = VideoPlayer(ventana_cilindrico, "cilindrico.mp4")
    videoplayer.pack(expand=True, fill="both")
    videoplayer.play()
###################################### AGARRE ESFÉRICO #############################################################
def evaluar_esferico(x, mano, evaluacion):
    print(x)
    if mano == "izquierda":
        angulos = x.values()
        if all(-3 <= angulo <= 3 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,0, x)
            return 0
        elif all(3 <= angulo <= 15 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,1, x)
            return 1
        elif all( 15 <= angulo <= 30 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,2, x)
            return 2
        else:
            return None
        pass
    elif mano == "derecha":
        angulos = x.values()
        # Logica para evaluar la mano derecha
        if all(-3 <= angulo <= 3 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,0, x)
            return 0
        elif all(3<= angulo <= 15 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,1, x)
            return 1
        elif all( 15 <= angulo <= 30 for angulo in angulos):
            agregar_resultado_mano(evaluacion, mano,2, x)
            return 2
        else:
            return None
        pass


def verificar_posicion_iniciales(mano):
    valores = leer_angulo(mano)
    print(valores)
    if all(-3 <= valor <= 3 for valor in valores.values()):
        return True
    return False

# Nueva función para realizar la evaluación de flexión después de la espera
def realizar_evaluaciones(mano, evaluacion):

    resultado = evaluar_esferico(leer_angulo(mano), mano, evaluacion)
    
    if resultado is not None:
        resultado_label.config(text=f"Resultado del Agarre Esférico: {resultado}")
    else:
        resultado_label.config(text="Valores fuera de rango.")

def iniciar_evaluacion_esferico(mano,ventana, evaluacion):
    if verificar_posicion_iniciales(mano):
        resultado_label.config(text="Realice el movimiento de Agarre Esférico.")
        # Esperar 3 segundos antes de realizar la evaluación
        ventana.after(5000, lambda: realizar_evaluaciones(mano, evaluacion))
    else:
        resultado_label.config(text="Por favor, ajusta la posición inicial.")
def abrir_agarre_esferico():
    ventana_esferico = tk.Toplevel()
    ventana_esferico.title("Agarre Esférico")
    tk.Label(ventana_esferico, text="Selecciona la mano:").pack()
    mano_var = tk.StringVar(value="izquierda")
    tk.Radiobutton(ventana_esferico, text="Izquierda", variable=mano_var, value="izquierda").pack()
    tk.Radiobutton(ventana_esferico, text="Derecha", variable=mano_var, value="derecha").pack()

    tk.Label(ventana_esferico, text="Nombre de la evaluación:").pack()
    entrada_evaluacion = tk.Entry(ventana_esferico)
    entrada_evaluacion.pack()
    iniciar_btn = tk.Button(ventana_esferico, text="Iniciar Evaluación",
                            command=lambda: iniciar_evaluacion_esferico(mano_var.get(), ventana_esferico, entrada_evaluacion.get()))
    iniciar_btn.pack()
    
    global resultado_label
    resultado_label = tk.Label(ventana_esferico, text="")
    resultado_label.pack()

    video_label = tk.Label(ventana_esferico, text="Video de referencia:")
    video_label.pack()
    
    videoplayer = VideoPlayer(ventana_esferico, "esferico.mp4")
    videoplayer.pack(expand=True, fill="both")
    videoplayer.play()
###################################### FMA MUÑECA #############################################################
# Función para cambiar a la ventana de evaluación FMA-Muñeca
def abrir_fma_muneca():
    ventana_fma_muneca = tk.Toplevel()
    ventana_fma_muneca.title("Evaluación FMA-Muñeca")

    
    # Creamos el widget Text y lo configuramos para que sea de solo lectura
    texto_descripcion = tk.Text(ventana_fma_muneca, wrap=tk.WORD, height=13, width=50)
    texto_descripcion.grid(row=0, column=0, columnspan=2)

    # Descripción de la prueba
    descripcion = ("Se va a llevar a cabo la evaluación Fugl-Meyer de la muñeca,\n "
                   "esta prueba consta de 5 movimientos: \n "
                   "1. Estabilidad de la muñeca con codo a 90 grados\n"
                   "2. Flexión y Extensión con codo a 90 grados\n "
                   "3. Estabilidad de la muñeca con codo a 0 grados\n"
                   "4. Flexión y Extensión con codo a 0 grados\n "
                   "5. circunducción\n "
                   "Cada uno de los movimientos se clasifica según la siguiente escala:\n"
                   "0- no puede desarrollar el movimiento\n "
                   "1- desarrolla el movimiento de forma parcia\n "
                   "2- desarrolla el movimiento de forma completa")
    # Insertamos el texto en el widget Text
    texto_descripcion.insert(tk.END, descripcion)

    # Centramos el texto utilizando una etiqueta de formato
    texto_descripcion.tag_configure("center", justify="center")
    texto_descripcion.tag_add("center", "1.0", "end")

    # Hacemos que el widget Text sea de solo lectura
    texto_descripcion.config(state=tk.DISABLED)

    # Botones para cada movimiento
    movimientos = [("1. Estabilidad 90°", 6, 0, abrir_Estabilidad_90), ("2. Flexión y Extensión 90°", 6, 1,abrir_flexion_90), ("3. Estabilidad 0°", 7, 0,abrir_Estabilidad_0 ),
                   ("4. Flexión y Extensión 0°", 7, 1, abrir_flexion_0), ("5. Circunducción", 9, 0,  abrir_Circunduccion)]

    button_width = 20  # Ancho deseado de los botones
    for movimiento, fila, columna, ventana in movimientos[:-1]:
        tk.Button(ventana_fma_muneca, text=movimiento, command = ventana, width=button_width).grid(row=fila, column=columna, padx=5, pady=5)

    # Colocamos el botón 'Agarre Esférico' en el centro de la tercera fila
    tk.Button(ventana_fma_muneca, text="5. Circunducción",command= abrir_Circunduccion, width=button_width).grid(row=9, column=0, columnspan=2, padx=5, pady=5)

    # Botón para volver a la ventana de inicio
    tk.Button(ventana_fma_muneca, text="Volver", command=ventana_fma_muneca.destroy, width=button_width).grid(row=10, column=0, columnspan=2, padx=5, pady=5)
###################################### ESTABILIDAD CON CODO A 90°#############################################################
def evaluar_Estabilidad9(x, mano, evaluacion):
    print(x)
    roll = x.pop('roll')
    if mano == "izquierda":
        if -3 <= roll <= 0 :
            print(roll)
            a = {'roll': roll}
            agregar_resultado_muneca(evaluacion, mano,0, a)
            return 0
        elif  -15 <= roll <= -3:
            print(roll)
            a = {'roll': roll}
            agregar_resultado_muneca(evaluacion, mano,1, a)
            return 1
        elif -30 <= roll <= -15:
            print(roll)
            a = {'roll': roll}
            agregar_resultado_muneca(evaluacion, mano,2, a)
            return 2
        else:
            print(roll)
            return None
        pass
    elif mano == "derecha":
        if -3 <= roll <= 0 :
            print(roll)
            a = {'roll': roll}
            agregar_resultado_muneca(evaluacion, mano,0, a)
            return 0
        elif  -15 <= roll <= -3:
            print(roll)
            a = {'roll': roll}
            agregar_resultado_muneca(evaluacion, mano,1, a)
            return 1
        elif -30 <= roll <= -15:
            print(roll)
            a = {'roll': roll}
            agregar_resultado_muneca(evaluacion, mano,2, a)
            return 2
        else:
            print(roll)
            return None
        pass

def verificar_posicion_inicialE9(mano):
    valores = mpu9250()
    roll = valores.pop('roll')
    print(roll)
    if -3<= roll <= 3 :
        return True
    return False

# Nueva función para realizar la evaluación de flexión después de la espera
def realizar_evaluacionE9(mano, evaluacion):

    resultado = evaluar_Estabilidad9(mpu9250(), mano, evaluacion)
    
    if resultado is not None:
        resultado_label.config(text=f"Resultado de la Estalibidad con codo a 90°: {resultado}")
    else:
        resultado_label.config(text="Valores fuera de rango.")

def iniciar_evaluacion_Estabilidad9(mano,ventana, evaluacion):
    if verificar_posicion_inicialE9(mano):
        resultado_label.config(text="Realice el movimiento de Estabilidad con codo a 90°.")
        # Esperar 3 segundos antes de realizar la evaluación
        ventana.after(5000, lambda: realizar_evaluacionE9(mano, evaluacion))
    else:
        resultado_label.config(text="Por favor, ajusta la posición inicial.")
def abrir_Estabilidad_90():
    ventana_Estabilidad9 = tk.Toplevel()
    ventana_Estabilidad9.title("Estabilidad con codo a 90°")
    tk.Label(ventana_Estabilidad9, text="Selecciona la mano:").pack()
    mano_var = tk.StringVar(value="izquierda")
    tk.Radiobutton(ventana_Estabilidad9, text="Izquierda", variable=mano_var, value="izquierda").pack()
    tk.Radiobutton(ventana_Estabilidad9, text="Derecha", variable=mano_var, value="derecha").pack()

    tk.Label(ventana_Estabilidad9, text="Nombre de la evaluación:").pack()
    entrada_evaluacion = tk.Entry(ventana_Estabilidad9)
    entrada_evaluacion.pack()

    iniciar_btn = tk.Button(ventana_Estabilidad9, text="Iniciar Evaluación",
                            command=lambda: iniciar_evaluacion_Estabilidad9(mano_var.get(), ventana_Estabilidad9, entrada_evaluacion.get()))
    iniciar_btn.pack()
    
    global resultado_label
    resultado_label = tk.Label(ventana_Estabilidad9, text="")
    resultado_label.pack()

    video_label = tk.Label(ventana_Estabilidad9, text="Video de referencia:")
    video_label.pack()
    
    videoplayer = VideoPlayer(ventana_Estabilidad9, "Estabilidad90.mp4")
    videoplayer.pack(expand=True, fill="both")
    videoplayer.play()
###################################### FLEXIÓN CON CODO A 90°#############################################################
def evaluar_flexion9(x, mano, evaluacion):
    print(x)
    rolle = x.pop('roll')
    if mano == "izquierda":
        if -3 <= rolle <= 3 :
            y = mpu9250()
            rollf = y.pop('roll')
            if -3 <= rollf <= 3:
                print(rolle, rollf)
                a = {'rolle': rolle, 'rollf': rollf}
                agregar_resultado_muneca(evaluacion, mano,0, a)
                return 0
        elif  -40 <= rolle <= -3:
            y = mpu9250()
            rollf = y.pop('roll')
            if 3 <= rollf <= 40:
                print(rolle, rollf)
                a = {'rolle': rolle, 'rollf': rollf}
                agregar_resultado_muneca(evaluacion, mano,1, a)
                return 1
        elif -80 <= rolle <= -40:
            y = mpu9250()
            rollf = y.pop('roll')
            if 40 <= rollf <= 80:
                print(rolle, rollf)
                a = {'rolle': rolle, 'rollf': rollf}
                agregar_resultado_muneca(evaluacion, mano,2, a)
                return 2
        else:
            return None
    elif mano == "derecha":
        if -3 <= rolle <= 3 :
            y = mpu9250()
            rollf = y.pop('roll')
            if -3 <= rollf <= 3:
                print(rolle, rollf)
                a = {'rolle': rolle, 'rollf': rollf}
                agregar_resultado_muneca(evaluacion, mano,0, a)
                return 0
        elif  -40 <= rolle <= -3:
            y = mpu9250()
            rollf = y.pop('roll')
            if 3 <= rollf <= 40:
                print(rolle, rollf)
                a = {'rolle': rolle, 'rollf': rollf}
                agregar_resultado_muneca(evaluacion, mano,1, a)
                return 1
        elif -80 <= rolle <= -40:
            y = mpu9250()
            rollf = y.pop('roll')
            if 40 <= rollf <= 80:
                print(rolle, rollf)
                a = {'rolle': rolle, 'rollf': rollf}
                agregar_resultado_muneca(evaluacion, mano,2, a)
                return 2
        else:
            return None

def verificar_posicion_inicialf9(mano):
    valores = mpu9250()
    roll = valores.pop('roll')
    print(roll)
    if -3 <= roll <= 3 :
        return True
    return False

def realizar_evaluacionf9(mano, evaluacion):
    resultado = evaluar_flexion9(mpu9250(), mano, evaluacion)
    if resultado is not None:
        resultado_label.config(text=f"Resultado de la Flexión y Extensión con codo a 90°: {resultado}")
    else:
        resultado_label.config(text="Valores fuera de rango.")

def iniciar_evaluacion_flexion9(mano, ventana, evaluacion):
    if verificar_posicion_inicialf9(mano):
        resultado_label.config(text="Realice el movimiento de Flexión y Extensión con codo a 90°.")
        ventana.after(5000, lambda: realizar_evaluacionf9(mano, evaluacion))
    else:
        resultado_label.config(text="Por favor, ajusta la posición inicial.")

def abrir_flexion_90():
    ventana_flexion9 = tk.Toplevel()
    ventana_flexion9.title("Flexión y Extensión con codo a 90°")
    tk.Label(ventana_flexion9, text="Selecciona la mano:").pack()
    mano_var = tk.StringVar(value="izquierda")
    tk.Radiobutton(ventana_flexion9, text="Izquierda", variable=mano_var, value="izquierda").pack()
    tk.Radiobutton(ventana_flexion9, text="Derecha", variable=mano_var, value="derecha").pack()

    tk.Label(ventana_flexion9, text="Nombre de la evaluación:").pack()
    entrada_evaluacion = tk.Entry(ventana_flexion9)
    entrada_evaluacion.pack()
    
    iniciar_btn = tk.Button(ventana_flexion9, text="Iniciar Evaluación",
                            command=lambda: iniciar_evaluacion_flexion9(mano_var.get(), ventana_flexion9, entrada_evaluacion.get()))
    iniciar_btn.pack()
    
    global resultado_label
    resultado_label = tk.Label(ventana_flexion9, text="")
    resultado_label.pack()

    video_label = tk.Label(ventana_flexion9, text="Video de referencia:")
    video_label.pack()
    
    videoplayer = VideoPlayer(ventana_flexion9, "eyf90-1.mp4")
    videoplayer.pack(expand=True, fill="both")
    videoplayer.play()
###################################### ESTABILIDAD CON CODO A 0°#############################################################
def evaluar_Estabilidad0(x, mano, evaluacion):
    print(x)
    roll = x.pop('roll')
    if mano == "izquierda":
        if -3 <= roll <= 3 :
            print(roll)
            a = {'roll' : roll}
            agregar_resultado_muneca(evaluacion, mano,0, a)
            return 0
        elif  -15 <= roll <= -3:
            print(roll)
            a = {'roll' : roll}
            agregar_resultado_muneca(evaluacion, mano,1, a)
            return 1
        elif -30 <= roll <= -15:
            print(roll)
            a = {'roll' : roll}
            agregar_resultado_muneca(evaluacion, mano,2, a)
            return 2
        else:
            print(roll)
            return None
        pass
    elif mano == "derecha":
        if -3 <= roll <= 3 :
            print(roll)
            a = {'roll' : roll}
            agregar_resultado_muneca(evaluacion, mano,0, a)
            return 0
        elif  -15 <= roll <= -3:
            print(roll)
            a ={'roll' : roll}
            agregar_resultado_muneca(evaluacion, mano,1, a)
            return 1
        elif -30 <= roll <= -15:
            print(roll)
            a ={'roll' : roll}
            agregar_resultado_muneca(evaluacion, mano,2, a)
            return 2
        else:
            print(roll)
            return None
        pass

def verificar_posicion_inicialE0(mano):
    valores = mpu9250()
    roll = valores.pop('roll')
    print(roll)
    if -3 <= roll <= 3 :
        return True
    return False

# Nueva función para realizar la evaluación de flexión después de la espera
def realizar_evaluacionE0(mano, evaluacion):

    resultado = evaluar_Estabilidad0(mpu9250(), mano, evaluacion)
    
    if resultado is not None:
        resultado_label.config(text=f"Resultado de la Estalibidad con codo a 0°: {resultado}")
    else:
        resultado_label.config(text="Valores fuera de rango.")

def iniciar_evaluacion_Estabilidad0(mano,ventana, evaluacion):
    if verificar_posicion_inicialE0(mano):
        resultado_label.config(text="Realice el movimiento de Estabilidad con codo a 0°.")
        # Esperar 3 segundos antes de realizar la evaluación
        ventana.after(5000, lambda: realizar_evaluacionE0(mano ,evaluacion))
    else:
        resultado_label.config(text="Por favor, ajusta la posición inicial.")
def abrir_Estabilidad_0():
    ventana_Estabilidad0 = tk.Toplevel()
    ventana_Estabilidad0.title("Estabilidad con codo a 0°")
    tk.Label(ventana_Estabilidad0, text="Selecciona la mano:").pack()
    mano_var = tk.StringVar(value="izquierda")
    tk.Radiobutton(ventana_Estabilidad0, text="Izquierda", variable=mano_var, value="izquierda").pack()
    tk.Radiobutton(ventana_Estabilidad0, text="Derecha", variable=mano_var, value="derecha").pack()

    tk.Label(ventana_Estabilidad0, text="Nombre de la evaluación:").pack()
    entrada_evaluacion = tk.Entry(ventana_Estabilidad0)
    entrada_evaluacion.pack()
    
    iniciar_btn = tk.Button(ventana_Estabilidad0, text="Iniciar Evaluación",
                            command=lambda: iniciar_evaluacion_Estabilidad0(mano_var.get(), ventana_Estabilidad0, entrada_evaluacion.get()))
    iniciar_btn.pack()
    
    global resultado_label
    resultado_label = tk.Label(ventana_Estabilidad0, text="")
    resultado_label.pack()

    video_label = tk.Label(ventana_Estabilidad0, text="Video de referencia:")
    video_label.pack()
    
    videoplayer = VideoPlayer(ventana_Estabilidad0, "estabilidad0-1.mp4")
    videoplayer.pack(expand=True, fill="both")
    videoplayer.play()
###################################### FLEXIÓN CON CODO A 90°#############################################################
def evaluar_flexion0(x, mano, evaluacion):
    print(x)
    rolle = x.pop('roll')
    if mano == "izquierda":
        if -3 <= rolle <= 3 :
            y = mpu9250()
            rollf = y.pop('roll')
            if -3 <= rollf <= 3:
                print(rolle, rollf)
                a = {'rolle': rolle, 'rollf': rollf}
                agregar_resultado_muneca(evaluacion, mano,0, a)
                return 0
        elif  -40 <= rolle <= -3:
            y = mpu9250()
            rollf = y.pop('roll')
            if 3 <= rollf <= 40:
                print(rolle, rollf)
                a = {'rolle': rolle, 'rollf': rollf}
                agregar_resultado_muneca(evaluacion, mano,1, a)
                return 1
        elif -80 <= rolle <= -40:
            y = mpu9250()
            rollf = y.pop('roll')
            if 40 <= rollf <= 80:
                print(rolle, rollf)
                a = {'rolle': rolle, 'rollf': rollf}
                agregar_resultado_muneca(evaluacion, mano,2, a)
                return 2
        else:
            return None

    elif mano == "derecha":
        if -3 <= rolle <= 3 :
            y = mpu9250()
            rollf = y.pop('roll')
            if -3 <= rollf <= 3:
                print(rolle, rollf)
                a = {'rolle': rolle, 'rollf': rollf}
                agregar_resultado_muneca(evaluacion, mano,0, a)
                return 0
        elif  -40 <= rolle <= -3:
            y = mpu9250()
            rollf = y.pop('roll')
            if 3 <= rollf <= 40:
                print(rolle, rollf)
                a = {'rolle': rolle, 'rollf': rollf}
                agregar_resultado_muneca(evaluacion, mano,1, a)
                return 1
        elif -80 <= rolle <= -40:
            y = mpu9250()
            rollf = y.pop('roll')
            if 40 <= rollf <= 80:
                print(rolle, rollf)
                a = {'rolle': rolle, 'rollf': rollf}
                agregar_resultado_muneca(evaluacion, mano,2, a)
                return 2
        else:
            return None


def verificar_posicion_inicialf0(mano):
    valores = mpu9250()
    roll = valores.pop('roll')
    print(roll)
    if -3 <= roll <= 3 :
        return True
    return False

def realizar_evaluacionf0(mano,evaluacion):
    resultado = evaluar_flexion0(mpu9250(), mano,evaluacion)
    if resultado is not None:
        resultado_label.config(text=f"Resultado de la Flexión y Extensión con codo a 0°: {resultado}")
    else:
        resultado_label.config(text="Valores fuera de rango.")

def iniciar_evaluacion_flexion0(mano, ventana, evaluacion):
    if verificar_posicion_inicialf0(mano):
        resultado_label.config(text="Realice el movimiento de Flexión y Extensión con codo a 0°.")
        ventana.after(5000, lambda: realizar_evaluacionf0(mano, evaluacion))
    else:
        resultado_label.config(text="Por favor, ajusta la posición inicial.")
def abrir_flexion_0():
    ventana_flexion0 = tk.Toplevel()
    ventana_flexion0.title("Flexión y Extensión con codo a 0°")
    tk.Label(ventana_flexion0, text="Selecciona la mano:").pack()
    mano_var = tk.StringVar(value="izquierda")
    tk.Radiobutton(ventana_flexion0, text="Izquierda", variable=mano_var, value="izquierda").pack()
    tk.Radiobutton(ventana_flexion0, text="Derecha", variable=mano_var, value="derecha").pack()

    tk.Label(ventana_flexion0, text="Nombre de la evaluación:").pack()
    entrada_evaluacion = tk.Entry(ventana_flexion0)
    entrada_evaluacion.pack()
    
    iniciar_btn = tk.Button(ventana_flexion0, text="Iniciar Evaluación",
                            command=lambda: iniciar_evaluacion_flexion0(mano_var.get(), ventana_flexion0, entrada_evaluacion.get()))
    iniciar_btn.pack()
    
    global resultado_label
    resultado_label = tk.Label(ventana_flexion0, text="")
    resultado_label.pack()

    video_label = tk.Label(ventana_flexion0, text="Video de referencia:")
    video_label.pack()
    
    videoplayer = VideoPlayer(ventana_flexion0, "eyf0.mp4")
    videoplayer.pack(expand=True, fill="both")
    videoplayer.play()
###################################### CIRCUNDUCCIÓN #############################################################
def evaluar_Circunduccion(x, mano,evaluacion):
    print(x)
    yawad = x.pop('yaw')
    rolle = x.pop('roll')
    print(yawad)
    if mano == "izquierda":
        if -3 <= yawad <= 3 :
            y = mpu9250()
            yawabd = y.pop('yaw')
            print(yawabd)
            if -3 <= yawabd <= 3:
                g = mpu9250()
                rolle = g.pop('roll')
                print(rolle)
                if -3 <= rolle <= 3 :
                    m = mpu9250()
                    rollf = m.pop('roll')
                    print(rollf)
                    if -3 <= rollf <= 3:
                        print(yawad,yawabd,rolle,rollf)
                        a ={'yawad': yawad,'yawabd': yawabd,'rolle': rolle, 'rollf': rollf}
                        agregar_resultado_muneca(evaluacion, mano,0, a)
                        return 0
        elif  -40 <= rolle <= -3:
            y = mpu9250()
            rollf = y.pop('roll')
            if 3 <= rollf <= 40:
                print(rolle, rollf)
                a ={'rolle': rolle, 'rollf': rollf}
                agregar_resultado_muneca(evaluacion, mano,1, a)
                return 1
        elif -45 <= yawad <= -15:
            y = mpu9250()
            yawabd = y.pop('yaw')
            print(yawabd)
            if 3 <= yawabd <= 15:
                g = mpu9250()
                rolle = g.pop('roll')
                print(rolle)
                if -80 <= rolle <= -40 :
                    m = mpu9250()
                    rollf = m.pop('roll')
                    print(rollf)
                    if 40 <= rollf <= 80:
                        print(yawad,yawabd,rolle,rollf)
                        a ={'yawad': yawad,'yawabd': yawabd,'rolle': rolle, 'rollf': rollf}
                        agregar_resultado_muneca(evaluacion, mano,2, a)
                        return 2
        else:
            return None
    elif mano == "derecha":
        if -3 <= yawad <= 3 :
            y = mpu9250()
            yawabd = y.pop('yaw')
            print(yawabd)
            if -3 <= yawabd <= 3:
                g = mpu9250()
                rolle = g.pop('roll')
                print(rolle)
                if -3 <= rolle <= 3 :
                    m = mpu9250()
                    rollf = m.pop('roll')
                    print(rollf)
                    if -3 <= rollf <= 3:
                        print(yawad,yawabd,rolle,rollf)
                        a ={'yawad': yawad,'yawabd': yawabd,'rolle': rolle, 'rollf': rollf}
                        agregar_resultado_muneca(evaluacion, mano,0, a)
                        return 0
        elif  -40 <= rolle <= -3:
            y = mpu9250()
            rollf = y.pop('roll')
            if 3 <= rollf <= 40:
                print(rolle, rollf)
                a ={'rolle': rolle, 'rollf': rollf}
                agregar_resultado_muneca(evaluacion, mano,1, a)
                return 1
        elif 3 <= yawad <= 45:
            y = mpu9250()
            yawabd = y.pop('yaw')
            print(yawabd)
            if -15 <= yawabd <= 3:
                g = mpu9250()
                rolle = g.pop('roll')
                print(rolle)
                if -80 <= rolle <= -40 :
                    m = mpu9250()
                    rollf = m.pop('roll')
                    print(rollf)
                    if 40 <= rollf <= 80:
                        print(yawad,yawabd,rolle,rollf)
                        a ={'yawad': yawad,'yawabd': yawabd,'rolle': rolle, 'rollf': rollf}
                        agregar_resultado_muneca(evaluacion, mano,2, a)
                        return 2
        else:
            return None


def verificar_posicion_inicialcir(mano):
    valores = mpu9250()
    yaw = valores.pop('yaw')
    print(yaw)
    if -3 <= yaw <= 3 :
        return True
    return False

def realizar_evaluacioncir(mano,evaluacion):
    resultado = evaluar_Circunduccion(mpu9250(), mano,evaluacion)
    if resultado is not None:
        resultado_label.config(text=f"Resultado de la Cricunducción: {resultado}")
    else:
        resultado_label.config(text="Valores fuera de rango.")

def iniciar_evaluacion_Circunduccion(mano, ventana,evaluacion):
    if verificar_posicion_inicialcir(mano):
        resultado_label.config(text="Realice el movimiento de Circunducción.")
        ventana.after(5000, lambda: realizar_evaluacioncir(mano,evaluacion))
    else:
        resultado_label.config(text="Por favor, ajusta la posición inicial.")
def abrir_Circunduccion():
    ventana_Circunduccion = tk.Toplevel()
    ventana_Circunduccion.title("Circunducción")
    tk.Label(ventana_Circunduccion, text="Selecciona la mano:").pack()
    mano_var = tk.StringVar(value="izquierda")
    tk.Radiobutton(ventana_Circunduccion, text="Izquierda", variable=mano_var, value="izquierda").pack()
    tk.Radiobutton(ventana_Circunduccion, text="Derecha", variable=mano_var, value="derecha").pack()

    tk.Label(ventana_Circunduccion , text="Nombre de la evaluación:").pack()
    entrada_evaluacion = tk.Entry(ventana_Circunduccion)
    entrada_evaluacion.pack()
    
    iniciar_btn = tk.Button(ventana_Circunduccion, text="Iniciar Evaluación",
                            command=lambda: iniciar_evaluacion_Circunduccion(mano_var.get(), ventana_Circunduccion, entrada_evaluacion.get()))
    iniciar_btn.pack()
    
    global resultado_label
    resultado_label = tk.Label(ventana_Circunduccion, text="")
    resultado_label.pack()

    video_label = tk.Label(ventana_Circunduccion, text="Video de referencia:")
    video_label.pack()
    
    videoplayer = VideoPlayer(ventana_Circunduccion, "circunduccion.mp4")
    videoplayer.pack(expand=True, fill="both")
    videoplayer.play()
###################################### VENTANA DE RESULTADOS #############################################################
# Simulación de resultados de las evaluaciones (para propósitos de demostración)
# Cada evaluación tiene asociados dos resultados.

# Función para abrir la ventana de resultados
def abrir_resultados():
    ventana_resultados = tk.Toplevel()
    ventana_resultados.title("Resultados de Evaluaciones")
    
    tk.Label(ventana_resultados, text="Resultados de Evaluaciones Fugl-Meyer").pack()

    nombre_completo = entrada_nombre.get()
    cedula_identidad = entrada_cedula.get()

    # Mostrar datos del usuario
    tk.Label(ventana_resultados, text=f"Nombre Completo: {nombre_completo}").pack()
    tk.Label(ventana_resultados, text=f"Cédula de Identidad: {cedula_identidad}").pack()

    # Mostrar resultados de las evaluaciones
    tk.Label(ventana_resultados, text="Resultados de la Mano:").pack()
    for key, value in resultados_mano.items():
        tk.Label(ventana_resultados, text=f"{key}: {value[0]}, {value[1]}, {value[2]}").pack()

    tk.Label(ventana_resultados, text="Resultados de la Muñeca:").pack()
    for key, value in resultados_muneca.items():
        tk.Label(ventana_resultados, text=f"{key}: {value[0]}, {value[1]},{value[2]}").pack()

    # Botón para exportar a Excel
    exportar_btn = tk.Button(ventana_resultados, text="Exportar a Excel", command=lambda: exportar_a_excel(nombre_completo, cedula_identidad))
    exportar_btn.pack()
###################################### EXPORTAR A EXCEL #############################################################

# Función para exportar los resultados a Excel

# Función para ajustar el ancho de las columnas
def exportar_a_excel(nombre, cedula):
    # Recolectar datos de las evaluaciones en una estructura adecuada para pandas
    datos_mano = []
    datos_muneca = []

    # Añadir los resultados de mano
    for key, value in resultados_mano.items():
        datos_mano.append({
            "Nombre Completo": nombre,
            "Cédula de Identidad": cedula,
            "Evaluación": key,
            "Mano": value[0],
            "Puntaje": value[1],
            "Ángulos": value[2]
        })

    # Añadir los resultados de muñeca
    for key, value in resultados_muneca.items():
        datos_muneca.append({
            "Nombre Completo": nombre,
            "Cédula de Identidad": cedula,
            "Evaluación": key,
            "Mano": value[0],
            "Puntaje": value[1],
            "Ángulos": value[2]
        })

    df_mano = pd.DataFrame(datos_mano)
    df_muneca = pd.DataFrame(datos_muneca)

    file_path = "resultados_evaluaciones1.xlsx"

    # Verificar si el archivo ya existe y es válido
    if os.path.exists(file_path):
        try:
            book = load_workbook(file_path)
        except (InvalidFileException, zipfile.BadZipFile):
            messagebox.showerror("Error", "El archivo de resultados existente está dañado. Se creará un nuevo archivo.")
            book = None
    else:
        book = None

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        if not df_mano.empty:
            sumas_mano = df_mano.groupby("Mano")["Puntaje"].sum().to_dict()
            total_derecha = sumas_mano.get("derecha", 0)
            total_izquierda = sumas_mano.get("izquierda", 0)
            df_mano["Puntaje Total Mano Derecha"] = total_derecha
            df_mano["Puntaje Total Mano Izquierda"] = total_izquierda
            df_mano.to_excel(writer, sheet_name="Mano", index=False)
        
        if not df_muneca.empty:
            sumas_muneca = df_muneca.groupby("Mano")["Puntaje"].sum().to_dict()
            total_derecha = sumas_muneca.get("derecha", 0)
            total_izquierda = sumas_muneca.get("izquierda", 0)
            df_muneca["Puntaje Total Muñeca Derecha"] = total_derecha
            df_muneca["Puntaje Total Muñeca Izquierda"] = total_izquierda
            df_muneca.to_excel(writer, sheet_name="Muñeca", index=False)

    if book:
        writer.book = book
        writer.save()
    else:
        book = load_workbook(file_path)

    if "Mano" in book.sheetnames:
        sheet_mano = book["Mano"]
        ajustar_ancho_columnas(sheet_mano)
        
    if "Muñeca" in book.sheetnames:
        sheet_muneca = book["Muñeca"]
        ajustar_ancho_columnas(sheet_muneca)
    
    book.save(file_path)

    messagebox.showinfo("Exportación Exitosa", "Los resultados se han exportado a resultados_evaluaciones1.xlsx")

# Otras funciones necesarias para el ajuste de ancho y totales
def ajustar_ancho_columnas(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width
        
def agregar_totales(sheet, columna_puntaje):
    max_row = sheet.max_row + 1
    total_formula = f"=SUM({columna_puntaje}2:{columna_puntaje}{sheet.max_row})"
    sheet[f"{columna_puntaje}{max_row}"] = "Total"
    sheet[f"{columna_puntaje}{max_row + 1}"] = total_formula
    sheet[f"{columna_puntaje}{max_row + 1}"].alignment = Alignment(horizontal='center', vertical='center')
def guardar_resultados():
    nombre = entrada_nombre.get()
    cedula = entrada_cedula.get()

    if not nombre or not cedula:
        messagebox.showerror("Error", "Por favor, ingrese el nombre completo y la cédula de identidad.")
        return

    exportar_a_excel(nombre, cedula)
###################################### VENTANA PRINCIPAL #############################################################

# Ventana principal
ventana_principal = tk.Tk()
ventana_principal.title("Registro de Usuario y Evaluaciones Fugl-Meyer")

# Entrada de datos del usuario
tk.Label(ventana_principal, text="Nombre Completo:").pack()
entrada_nombre = tk.Entry(ventana_principal)
entrada_nombre.pack()

tk.Label(ventana_principal, text="Cédula de Identidad:").pack()
entrada_cedula = tk.Entry(ventana_principal)
entrada_cedula.pack()

button_width = 30  # Ancho deseado de los botones
# Botones para iniciar evaluaciones
tk.Button(ventana_principal, text="Iniciar Evaluación Fugl-Meyer Mano", command=abrir_fma_mano, width=button_width).pack()
tk.Button(ventana_principal, text="Iniciar Evaluación Fugl-Meyer Muñeca", command=abrir_fma_muneca, width=button_width).pack()

# Botón para abrir resultados
tk.Button(ventana_principal, text="Ver Resultados y Exportar a Excel", command=abrir_resultados, width=button_width).pack()

# Cargar imágenes
imagen1 = Image.open("poini.jpeg").resize((300, 400))
imagen2 = Image.open("poinapo.jpeg").resize((300, 400))


# Convertir imágenes a formato compatible con tkinter
imagen1_tk = ImageTk.PhotoImage(imagen1)
imagen2_tk = ImageTk.PhotoImage(imagen2)

# Mostrar imágenes y títulos
frame_imagenes = tk.Frame(ventana_principal)
frame_imagenes.pack()

tk.Label(frame_imagenes, image=imagen1_tk).grid(row=0, column=0)
tk.Label(frame_imagenes, text="Posición inicial sin apoyo").grid(row=1, column=0)

tk.Label(frame_imagenes, image=imagen2_tk).grid(row=0, column=1)
tk.Label(frame_imagenes, text="Posición inicial con apoyo").grid(row=1, column=1)


ventana_principal.mainloop()
